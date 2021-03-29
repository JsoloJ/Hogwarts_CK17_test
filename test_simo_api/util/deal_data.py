from openpyxl import load_workbook
from test_api.util.log import logger

#这里的openpyxl只能处理新版的excel,及后缀为xlsx
# def read_data(path,sheetname,max_r,min_r=1):
def read_data(path,sheetname):
# def read_data():
    # path ='../data/test.xlsx'
#     logger.info('读取测试数据，如果你需要定义数据的开头和结束行，你可以通过传参实现')
    wk = load_workbook(path) #加载已经存在的excel
    # wk_sheet = wk['del_mens']
    wk_sheet = wk[sheetname]
    testcase_data = []
    n = 1
    for x in range(1,len(tuple(wk_sheet.rows))+1):
    # for x in range(min_r,max_r +1):
        if x > 1:
            testcase_data1 = {}
            for y in range(1,10):
                    testcase_data1[wk_sheet.cell(row = 1,column =y).value] = wk_sheet.cell(row = x,column =y).value
            # print(testcase_data1)
            n = n + 1
            testcase_data1['nrow'] = n
            testcase_data.append(testcase_data1)
    print(testcase_data)
    # case_name = testcase_data
    wk.close()
    return testcase_data

# read_data()


#这里的openpyxl只能处理新版的excel,及后缀为xlsx
def write_data(path,sheetname,nrow,actual_results,msg):
# def deal_data():
    # path ='../data/test.xlsx'
    # logger.info('写入数据')
    wk = load_workbook(path) #加载已经存在的excel
    # wk_sheet = wk['data']
    wk_sheet = wk[sheetname]
    # 写入运行实际结果
    wk_sheet.cell(row=nrow,column=8,value=actual_results)
    #写入断言结果，可以自定义
    wk_sheet.cell(row=nrow,column=9,value=msg)
    wk.save(path)
    wk.close()